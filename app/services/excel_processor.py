import math
from numbers import Number

from openpyxl import load_workbook

from app.models.options import Action, ProcessingOptions, ProcessingResult, RoundingMode


def _is_numeric(value: object) -> bool:
    return isinstance(value, Number) and not isinstance(value, bool)


def _apply_rounding(value: float, mode: RoundingMode) -> float:
    if mode == RoundingMode.UP:
        return float(math.ceil(value))
    if mode == RoundingMode.DOWN:
        return float(math.floor(value))
    return value


def process_column_percentage(options: ProcessingOptions) -> ProcessingResult:
    workbook = load_workbook(options.input_path, data_only=False)
    sheet = workbook.active

    factor = 1 + (options.percentage / 100)
    if options.action == Action.DECREASE:
        factor = 1 - (options.percentage / 100)

    processed_cells = 0
    skipped_empty = 0
    skipped_formula = 0
    skipped_non_numeric = 0

    for row in range(1, sheet.max_row + 1):
        cell = sheet[f"{options.column}{row}"]

        if cell.value in (None, ""):
            skipped_empty += 1
            continue

        if cell.data_type == "f":
            skipped_formula += 1
            continue

        if not _is_numeric(cell.value):
            skipped_non_numeric += 1
            continue

        updated_value = float(cell.value) * factor

        if options.action == Action.DECREASE:
            updated_value = _apply_rounding(updated_value, options.rounding)

        cell.value = updated_value
        processed_cells += 1

    workbook.save(options.output_path)

    return ProcessingResult(
        sheet_name=sheet.title,
        processed_cells=processed_cells,
        skipped_empty=skipped_empty,
        skipped_formula=skipped_formula,
        skipped_non_numeric=skipped_non_numeric,
    )
